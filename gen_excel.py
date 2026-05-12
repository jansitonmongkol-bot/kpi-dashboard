#!/usr/bin/env python3
"""
gen_excel.py — สร้างรายงาน KPI ในรูปแบบ Excel (.xlsx) ด้วย openpyxl
Usage: python3 gen_excel.py <input.json> <output.xlsx>
"""
import sys, json
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel

def main():
    if len(sys.argv) < 3:
        print("Usage: gen_excel.py <input.json> <output.xlsx>")
        sys.exit(1)

    with open(sys.argv[1], encoding='utf-8') as f:
        d = json.load(f)

    outpath = sys.argv[2]
    rows    = d['rows']
    total   = d['total']

    # ── Colours ──────────────────────────────────────────────
    C_HEADER_BG   = "0F766E"   # teal
    C_HEADER_FG   = "FFFFFF"
    C_TITLE_BG    = "065F46"   # dark green
    C_PASS_BG     = "DCFCE7"   # light green
    C_FAIL_BG     = "FEE2E2"   # light red
    C_PASS_FG     = "15803D"
    C_FAIL_FG     = "B91C1C"
    C_STAT_BG     = "F0FDF4"
    C_ALT_ROW     = "F8FAFC"
    C_BORDER      = "CBD5E1"
    C_SUMMARY_BG  = "EFF6FF"

    def mk_fill(hex_str):
        return PatternFill("solid", fgColor=hex_str)

    def mk_border(all_sides="thin"):
        s = Side(style=all_sides, color=C_BORDER)
        return Border(left=s, right=s, top=s, bottom=s)

    def mk_font(bold=False, size=10, color="1E293B", name="Arial"):
        return Font(bold=bold, size=size, color=color, name=name)

    thin_border = mk_border()

    wb = Workbook()

    # ════════════════════════════════════════════════════════════
    #  SHEET 1: ข้อมูลรายละเอียด
    # ════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "รายละเอียด KPI"
    ws.sheet_view.showGridLines = False

    # ── Title block ──────────────────────────────────────────
    ws.merge_cells("A1:I1")
    ws["A1"] = f"รายงานข้อมูลตัวชี้วัด (KPI)"
    ws["A1"].font  = Font(bold=True, size=16, color=C_HEADER_FG, name="Arial")
    ws["A1"].fill  = mk_fill(C_TITLE_BG)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    meta = [
        ("A2", "รหัส KPI",      "B2", d["kpi_code"]),
        ("A3", "ชื่อตัวชี้วัด", "B3", d["kpi_name"]),
        ("A4", "เป้าหมาย",      "B4", f"{d['operator']} {d['target']}%"),
        ("A5", "วันที่ออกรายงาน","B5", d["date"]),
    ]
    ws.merge_cells("B2:I2")
    ws.merge_cells("B3:I3")
    ws.merge_cells("B4:I4")
    ws.merge_cells("B5:I5")

    for (lk, lv, vk, vv) in meta:
        ws[lk] = lv
        ws[lk].font = mk_font(bold=True, size=11)
        ws[lk].fill = mk_fill(C_STAT_BG)
        ws[lk].alignment = Alignment(horizontal="right", vertical="center", indent=1)
        ws[vk] = vv
        ws[vk].font = mk_font(size=11)
        ws[vk].fill = mk_fill("FFFFFF")
        ws[vk].alignment = Alignment(horizontal="left", vertical="center", indent=1)
        for cell in [ws[lk], ws[vk]]:
            cell.border = thin_border

    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 20

    # ── Summary Stats ────────────────────────────────────────
    ws["A7"] = "สรุปผลการดำเนินงาน"
    ws["A7"].font  = mk_font(bold=True, size=12, color=C_HEADER_FG)
    ws["A7"].fill  = mk_fill(C_HEADER_BG)
    ws.merge_cells("A7:I7")
    ws["A7"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[7].height = 28

    stats = [
        ("รายการทั้งหมด", total, ""),
        ("ผ่านเกณฑ์",      d["pass_count"], C_PASS_FG),
        ("ไม่ผ่านเกณฑ์",   d["fail_count"], C_FAIL_FG),
        ("อัตราผ่านเกณฑ์", f"{d['pass_rate']}%", ""),
        ("ค่าเฉลี่ย",       f"{d['avg_val']}%", ""),
        ("ค่าสูงสุด",       f"{d['max_val']}%", C_PASS_FG),
        ("ค่าต่ำสุด",       f"{d['min_val']}%", C_FAIL_FG),
    ]

    col = 1
    ws.row_dimensions[8].height = 18
    ws.row_dimensions[9].height = 24
    for label, value, fg in stats:
        lc = get_column_letter(col)
        vc = get_column_letter(col)
        ws[f"{lc}8"] = label
        ws[f"{lc}8"].font  = mk_font(bold=True, size=9, color="64748B")
        ws[f"{lc}8"].fill  = mk_fill(C_SUMMARY_BG)
        ws[f"{lc}8"].alignment = Alignment(horizontal="center")
        ws[f"{lc}8"].border = thin_border

        ws[f"{lc}9"] = value
        ws[f"{lc}9"].font  = Font(bold=True, size=13, color=fg if fg else "1E293B", name="Arial")
        ws[f"{lc}9"].fill  = mk_fill(C_SUMMARY_BG)
        ws[f"{lc}9"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"{lc}9"].border = thin_border
        col += 1

    # ── Data Table Header ─────────────────────────────────────
    DATA_ROW_START = 12
    ws[f"A{DATA_ROW_START-1}"] = "ข้อมูลรายละเอียด"
    ws[f"A{DATA_ROW_START-1}"].font = mk_font(bold=True, size=12, color=C_HEADER_FG)
    ws[f"A{DATA_ROW_START-1}"].fill = mk_fill(C_HEADER_BG)
    ws.merge_cells(f"A{DATA_ROW_START-1}:I{DATA_ROW_START-1}")
    ws[f"A{DATA_ROW_START-1}"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[DATA_ROW_START-1].height = 28

    # ── Detect row format (custom SQL vs default) ─────────────
    sample = rows[0] if rows else {}
    # Custom headers passed from PHP, or detect from row keys
    custom_headers = d.get('headers')  # list or None
    has_standard = 'period' in sample and 'actual' in sample and 'diff' in sample

    if custom_headers:
        # Use custom headers from PHP
        display_cols = [h for h in custom_headers if h not in ('#','ลำดับ','สถานะ','pass','actual','_no','_pass','_actual','no')]
        headers = ['#'] + display_cols + ['ผลงาน(%)', 'สถานะ']
        def get_row_vals(row, is_pass):
            vals = [row.get('no', row.get('_no', 0))]
            for h in display_cols:
                vals.append(row.get(h, ''))
            vals.append(float(row.get('actual', row.get('_actual', 0))))
            vals.append('บรรลุเป้าหมาย' if is_pass else 'ต่ำกว่าเป้าหมาย')
            return vals
    elif has_standard:
        # Standard default format
        headers = ['#', 'ช่วงเวลา', 'วันที่', 'ตัวเศษ', 'ตัวส่วน',
                   'ผลงาน (%)', 'เทียบเป้าหมาย (%)', 'สถานะ', 'หมายเหตุ']
        def get_row_vals(row, is_pass):
            return [
                row['no'], row['period'], row['date'],
                row['numerator'] if row['numerator'] != '' else None,
                row['denominator'] if row['denominator'] != '' else None,
                float(row['actual']),
                float(row.get('diff', float(row['actual']) - d['target'])),
                'บรรลุเป้าหมาย' if is_pass else 'ต่ำกว่าเป้าหมาย',
                row.get('note', ''),
            ]
    else:
        # Flexible: use all keys from row except internal ones
        skip = {'pass','_pass','_no','_actual','no'}
        row_keys = [k for k in sample.keys() if k not in skip]
        headers = ['#'] + row_keys + ['สถานะ']
        def get_row_vals(row, is_pass):
            vals = [row.get('no', row.get('_no', 0))]
            for k in row_keys:
                vals.append(row.get(k, ''))
            vals.append('บรรลุเป้าหมาย' if is_pass else 'ต่ำกว่าเป้าหมาย')
            return vals

    num_cols = len(headers)
    # Merge title rows to match new column count
    # (already merged A1:I1 above — fix if needed)

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=DATA_ROW_START, column=ci, value=h)
        cell.font      = Font(bold=True, size=10, color=C_HEADER_FG, name="Arial")
        cell.fill      = mk_fill(C_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin_border
    ws.row_dimensions[DATA_ROW_START].height = 22

    # ── Data Rows ────────────────────────────────────────────
    for ri, row in enumerate(rows):
        r       = DATA_ROW_START + 1 + ri
        is_pass = row.get('pass', row.get('_pass', False))
        is_alt  = ri % 2 == 1
        row_bg  = C_ALT_ROW if is_alt else "FFFFFF"
        vals    = get_row_vals(row, is_pass)
        status_ci = len(vals)  # last column index (1-based)
        # find actual value column index (look for float value)
        actual_ci = None
        for ci_scan, v in enumerate(vals, 1):
            try:
                float(v)
                actual_ci = ci_scan
            except (TypeError, ValueError):
                pass

        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.border    = thin_border
            cell.font      = mk_font(size=10)
            cell.fill      = mk_fill(row_bg)
            cell.alignment = Alignment(vertical="center", horizontal="center" if ci == 1 else "left", indent=1)

            # Last column = status
            if ci == status_ci:
                cell.font = Font(bold=True, size=10,
                                 color=C_PASS_FG if is_pass else C_FAIL_FG, name="Arial")
                cell.fill = mk_fill(C_PASS_BG if is_pass else C_FAIL_BG)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            # Numeric cells: center align
            elif isinstance(val, (int, float)):
                cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[r].height = 20

    # ── Summary row ──────────────────────────────────────────
    sumrow  = DATA_ROW_START + 1 + total
    last_c  = get_column_letter(num_cols)
    half_c  = max(1, num_cols // 2)
    merge_e = get_column_letter(half_c)

    ws.merge_cells(f"A{sumrow}:{merge_e}{sumrow}")
    ws[f"A{sumrow}"] = f"{d['pass_count']} ผ่านเกณฑ์ / {d['fail_count']} ไม่ผ่าน | อัตรา {d['pass_rate']}%"
    ws[f"A{sumrow}"].font      = mk_font(bold=True, size=10)
    ws[f"A{sumrow}"].fill      = mk_fill(C_STAT_BG)
    ws[f"A{sumrow}"].alignment = Alignment(horizontal="center", vertical="center")
    ws[f"A{sumrow}"].border    = thin_border
    ws.row_dimensions[sumrow].height = 22

    # fill rest of summary row
    for ci in range(half_c + 1, num_cols + 1):
        cell = ws.cell(row=sumrow, column=ci)
        cell.fill   = mk_fill(C_STAT_BG)
        cell.border = thin_border

    # ── Column widths ─────────────────────────────────────────
    default_w = 16
    col_widths_map = {1: 5}  # # column is narrow
    for i in range(1, num_cols + 1):
        w = col_widths_map.get(i, default_w)
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Freeze + Auto-filter ─────────────────────────────────
    last_col_letter = get_column_letter(num_cols)
    ws.auto_filter.ref = f"A{DATA_ROW_START}:{last_col_letter}{DATA_ROW_START+total}"
    ws.freeze_panes    = f"A{DATA_ROW_START+1}"

    # ════════════════════════════════════════════════════════════
    #  SHEET 2: กราฟข้อมูล
    # ════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("กราฟ & วิเคราะห์")
    ws2.sheet_view.showGridLines = False

    # Write raw data for chart (hidden columns)
    ws2["A1"] = "ช่วงเวลา"
    ws2["B1"] = "ผลงาน (%)"
    ws2["C1"] = f"เป้าหมาย ({d['operator']} {d['target']}%)"

    for i, row in enumerate(rows, 2):
        # Get label: use period, or first string-ish value
        lbl = row.get('period', row.get('ช่วงเวลา', ''))
        if not lbl:
            for v in row.values():
                if isinstance(v, str) and v: lbl = v; break
        # Get actual: use _actual, actual, or pass_rate col
        act = row.get('_actual', row.get('actual', 0))
        try: act = float(act)
        except: act = 0.0
        ws2.cell(row=i, column=1).value = lbl
        ws2.cell(row=i, column=2).value = act
        ws2.cell(row=i, column=3).value = float(d["target"])

    # Header styling for chart data
    for c in range(1, 4):
        cell = ws2.cell(row=1, column=c)
        cell.font  = Font(bold=True, size=10, color=C_HEADER_FG, name="Arial")
        cell.fill  = mk_fill(C_HEADER_BG)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    # ── Bar Chart ─────────────────────────────────────────────
    n = len(rows)
    bar = BarChart()
    bar.type    = "col"
    bar.grouping = "clustered"
    bar.title   = f"ผลงาน KPI {d['kpi_code']} รายช่วงเวลา"
    bar.y_axis.title = "ร้อยละ (%)"
    bar.x_axis.title = "ช่วงเวลา"
    bar.style   = 10
    bar.width   = 20
    bar.height  = 12

    data_ref   = Reference(ws2, min_col=2, min_row=1, max_row=n+1)
    target_ref = Reference(ws2, min_col=3, min_row=1, max_row=n+1)
    cats_ref   = Reference(ws2, min_col=1, min_row=2, max_row=n+1)

    bar.add_data(data_ref, titles_from_data=True)
    bar.set_categories(cats_ref)
    bar.series[0].graphicalProperties.solidFill = "0D9488"

    # Add target line as line series
    line_chart = LineChart()
    line_chart.add_data(target_ref, titles_from_data=True)
    line_chart.set_categories(cats_ref)
    line_chart.series[0].graphicalProperties.line.solidFill = "DC2626"
    line_chart.series[0].graphicalProperties.line.width = 20000  # 2pt
    line_chart.series[0].graphicalProperties.line.dashDot = "dash"

    bar += line_chart
    ws2.add_chart(bar, "E2")

    # ── Trend Line Chart ──────────────────────────────────────
    trend = LineChart()
    trend.title   = f"แนวโน้มผลงาน KPI {d['kpi_code']}"
    trend.y_axis.title = "ร้อยละ (%)"
    trend.x_axis.title = "ช่วงเวลา"
    trend.style   = 10
    trend.width   = 20
    trend.height  = 12

    trend.add_data(data_ref, titles_from_data=True)
    trend.add_data(target_ref, titles_from_data=True)
    trend.set_categories(cats_ref)

    trend.series[0].graphicalProperties.line.solidFill   = "2563EB"
    trend.series[0].graphicalProperties.line.width       = 25000
    trend.series[0].smooth = True

    trend.series[1].graphicalProperties.line.solidFill   = "DC2626"
    trend.series[1].graphicalProperties.line.width       = 18000
    trend.series[1].graphicalProperties.line.dashDot     = "dash"

    ws2.add_chart(trend, "E20")

    # Column widths for sheet2
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 18

    # ════════════════════════════════════════════════════════════
    #  SHEET 3: สรุปผล
    # ════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("สรุปผล")
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells("B2:F2")
    ws3["B2"] = "สรุปผลการดำเนินงานตัวชี้วัด"
    ws3["B2"].font  = Font(bold=True, size=18, color=C_TITLE_BG, name="Arial")
    ws3["B2"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[2].height = 40

    ws3.merge_cells("B3:F3")
    ws3["B3"] = d["kpi_name"]
    ws3["B3"].font  = Font(bold=True, size=13, color="475569", name="Arial")
    ws3["B3"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws3.row_dimensions[3].height = 35

    summary_items = [
        ("เป้าหมาย",              f"{d['operator']} {d['target']}%",  C_HEADER_BG, C_HEADER_FG),
        ("จำนวนข้อมูล",           f"{total} รายการ",                  "E0F2FE",    "0369A1"),
        ("ผ่านเกณฑ์",             f"{d['pass_count']} รายการ",        C_PASS_BG,   C_PASS_FG),
        ("ไม่ผ่านเกณฑ์",          f"{d['fail_count']} รายการ",        C_FAIL_BG,   C_FAIL_FG),
        ("อัตราผ่านเกณฑ์",        f"{d['pass_rate']}%",               "F0F9FF",    "0369A1"),
        ("ค่าเฉลี่ย",              f"{d['avg_val']}%",                 "F0FDF4",    "166534"),
        ("ค่าสูงสุด",              f"{d['max_val']}%",                 "DCFCE7",    C_PASS_FG),
        ("ค่าต่ำสุด",              f"{d['min_val']}%",                 "FEF2F2",    C_FAIL_FG),
    ]

    start_r = 5
    for ri, (label, value, bg, fg) in enumerate(summary_items):
        r = start_r + ri * 3
        ws3.merge_cells(f"B{r}:C{r+1}")
        ws3.merge_cells(f"D{r}:F{r+1}")

        ws3[f"B{r}"] = label
        ws3[f"B{r}"].font  = Font(bold=True, size=12, color="475569", name="Arial")
        ws3[f"B{r}"].fill  = mk_fill("F8FAFC")
        ws3[f"B{r}"].alignment = Alignment(horizontal="right", vertical="center", indent=2)
        ws3[f"B{r}"].border = thin_border

        ws3[f"D{r}"] = value
        ws3[f"D{r}"].font  = Font(bold=True, size=16, color=fg, name="Arial")
        ws3[f"D{r}"].fill  = mk_fill(bg)
        ws3[f"D{r}"].alignment = Alignment(horizontal="center", vertical="center")
        ws3[f"D{r}"].border = thin_border

        ws3.row_dimensions[r].height   = 18
        ws3.row_dimensions[r+1].height = 18

    for col in ["B","C","D","E","F"]:
        ws3.column_dimensions[col].width = 16

    # ── Save ─────────────────────────────────────────────────
    wb.save(outpath)
    print(f"OK: {outpath}")

if __name__ == "__main__":
    main()